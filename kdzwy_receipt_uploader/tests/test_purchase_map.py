from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook

PROJECT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT / "src"))

from kdzwy_receipt_uploader.purchase_map import build_purchase_map  # noqa: E402


def main() -> None:
    with TemporaryDirectory() as directory:
        root = Path(directory)
        source = root / "用途确认信息.xlsx"
        output = root / "maps" / "purchase_map.json"
        report = root / "maps" / "purchase_map.report.json"
        wb = Workbook()
        ws = wb.active
        ws.title = "发票"
        for column, title in {"E": "数电发票号码", "H": "开票日期", "J": "销售方名称", "K": "金额", "L": "税额"}.items():
            ws[f"{column}3"] = title
        ws["E4"] = "26112000002695439356"
        ws["H4"] = datetime(2026, 7, 31)
        ws["J4"] = "供应商A"
        ws["K4"] = 100
        ws["L4"] = 13
        ws["E5"] = "26112000002695439356"
        ws["H5"] = "2026/07/31"
        ws["J5"] = "供应商A"
        ws["K5"] = 50
        ws["L5"] = 6.5
        wb.save(source)
        wb.close()

        result = build_purchase_map(source, output, report)
        assert result["map"]["26112000002695439356"] == {
            "amount": 150.0,
            "taxAmount": 19.5,
            "totalAmount": 169.5,
            "date": "2026-07-31",
            "itemClass": "供应商",
            "supplierName": "供应商A",
            "supplierNameCandidates": ["供应商A"],
            "rowCount": 2,
        }
        report_payload = json.loads(report.read_text(encoding="utf-8"))
        assert report_payload["sourceSheet"] == "发票"
        assert report_payload["columns"] == {
            "invoiceCode": "E",
            "supplierName": "J",
            "amount": "K",
            "taxAmount": "L",
            "totalAmount": "K + L (逐行计算)",
            "date": "H",
        }
        assert json.loads(output.read_text(encoding="utf-8"))["26112000002695439356"]["supplierName"] == "供应商A"
        print("purchase_map 测试通过")


if __name__ == "__main__":
    main()

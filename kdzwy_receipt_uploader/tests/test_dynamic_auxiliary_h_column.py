from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook

PROJECT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT / "src"))

from kdzwy_receipt_uploader.month_config import MonthConfig
from kdzwy_receipt_uploader.receipt_generation import generate_receipts
from kdzwy_receipt_uploader.workflow import load_receipt
from kdzwy_receipt_uploader.sales_map import build_sales_map


def main() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        month = root / "weiyu" / "7月"
        (month / "sales1").mkdir(parents=True)
        (month / "sales1" / "dzfp_1001_customer.pdf").write_bytes(b"%PDF")
        (month / "month.conf").write_text("[processing]\ncompany=weiyu\nmonth=7月\n", encoding="utf-8")
        source = month / "收入成本表.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "信息汇总表"
        ws["D1"] = "发票号"
        ws["H1"] = "购买方名称"
        ws["I1"] = "开票日期"
        ws["Q1"] = "金额"
        ws["S1"] = "税额"
        ws["T1"] = "价税合计"
        ws["D2"] = "1001"
        ws["H2"] = "动态客户A"
        ws["I2"] = datetime(2026, 7, 1)
        ws["Q2"] = -100
        ws["S2"] = -13
        ws["T2"] = -113
        wb.save(source)
        sales_map = build_sales_map(source)["map"]
        assert sales_map["1001"]["customName"] == "动态客户A"
        cfg = MonthConfig.load(month / "month.conf")
        generate_receipts(
            month, cfg, month / "receipts", overwrite=True, folder_patterns=["sales*"],
            voucher_defaults={"group_id": "g", "group_name": "记", "user_name": "u", "itemClass": "客户"},
            entry_defaults=[{"account_number": "1122", "account_id": "a", "account_name": "应收账款", "dc": 1}],
            sales_map_values={"1001": {**sales_map["1001"], "auxiliaryItem": {"itemClass": "客户", "itemClassId": 1, "id": "live-c1", "number": "009", "name": "动态客户A"}}},
            template_config={"voucher_templates": [{"name": "客户", "when": {"itemClass": "客户"}, "summary": {"header": "收款", "body": "{invoiceCode}", "separator": " "}, "entries": [{"dc": 1, "accountSelector": {"number": "1122"}, "amountFrom": "sales_map.totalAmount", "auxiliary": {"field": "customId", "itemClassId": 1}}]}]},
        )
        payload = json.loads((month / "receipts" / "receipt_1001" / "receipt.json").read_text(encoding="utf-8"))
        entry = payload["voucher"]["entries"][0]
        assert entry["customId"] == "live-c1"
        assert entry["accountName"] == "动态客户A"
        assert entry["subjectAccountName"] == "应收账款"
        assert "customNumber" not in entry
        assert "customName" not in entry
        assert entry["auxiliaryExpected"] == {"itemClass": "客户", "id": "live-c1", "number": "009", "name": "动态客户A", "accountNumber": "1122", "subjectAccountName": "应收账款"}
        assert entry["amount"] == -113
        assert payload["voucher"]["date"] == "2026-07-01"
        print("H列动态辅助核算和负数金额测试通过")


if __name__ == "__main__":
    main()

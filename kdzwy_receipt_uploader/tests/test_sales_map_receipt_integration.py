from __future__ import annotations

import json
import sys
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

PROJECT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT / "src"))

from kdzwy_receipt_uploader.month_config import MonthConfig  # noqa: E402
from kdzwy_receipt_uploader.receipt_generation import generate_receipts  # noqa: E402
from kdzwy_receipt_uploader.sales_map import build_sales_map  # noqa: E402


def main() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        month = root / "weiyu" / "7月"
        (month / "sales1").mkdir(parents=True)
        pdf = month / "sales1" / "dzfp_1001_vendor.pdf"
        pdf.write_bytes(b"%PDF")
        source = month / "收入成本表.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "信息汇总表"
        ws["D1"] = "发票号"
        ws["D2"] = "1001"
        ws["Q2"] = 100
        ws["S2"] = 13
        ws["T2"] = 113
        ws["I2"] = datetime(2026, 7, 31)
        ws["D3"] = "1001"
        ws["Q3"] = 50
        ws["S3"] = 6.5
        ws["T3"] = 56.5
        ws["I3"] = datetime(2026, 7, 31)
        wb.save(source)
        sales_map_path = month / "maps" / "sales_map.json"
        result = build_sales_map(source, sales_map_path, month / "maps" / "sales_map.report.json")
        assert result["map"]["1001"]["amount"] == 150.0
        assert result["map"]["1001"]["taxAmount"] == 19.5
        assert result["map"]["1001"]["totalAmount"] == 169.5
        assert result["map"]["1001"]["date"] == "2026-07-31"
        assert result["map"]["1001"]["itemClass"] == "客户"
        assert result["map"]["1001"]["customName"] == ""
        assert result["map"]["1001"]["rowCount"] == 2
        (month / "weiyu_7月.conf").write_text("[processing]\ncompany=weiyu\nmonth=7月\n", encoding="utf-8")
        map_path = month / "maps" / "xlsx_pdf_map.json"
        map_path.parent.mkdir(parents=True, exist_ok=True)
        map_path.write_text(json.dumps({"1001": str(pdf)}), encoding="utf-8")
        cfg = MonthConfig.load(month / "weiyu_7月.conf")
        report = generate_receipts(
            month,
            cfg,
            month / "receipts",
            overwrite=True,
            map_file=map_path,
            folder_patterns=["sales*"],
            voucher_defaults={"group_id": "g1", "group_name": "记", "user_name": "用户"},
            entry_defaults=[{"line_no": 1, "dc": 1}, {"line_no": 2, "dc": -1}],
            sales_map_values=result["map"],
            template_config={
                "voucher_templates": [{
                    "name": "default",
                    "when": {},
                    "summary": {"header": "", "body": "", "separator": ""},
                    "entries": [{"dc": 1}, {"dc": -1}, {"dc": 1}],
                }]
            },
        )
        assert report["summary"]["generatedCount"] == 1
        payload = json.loads((month / "receipts" / "receipt_1001" / "receipt.json").read_text(encoding="utf-8"))
        assert payload["voucher"]["date"] == "2026-07-31"
        assert payload["voucher"]["totalAmount"] == 169.5
        assert len(payload["voucher"]["entries"]) == 3
        print("sales_map 与 receipt 模板集成测试通过")


if __name__ == "__main__":
    main()

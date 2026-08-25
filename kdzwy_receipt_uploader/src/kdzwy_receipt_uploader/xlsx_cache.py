"""Process-local cache for read-only Excel workbooks."""
from __future__ import annotations

import atexit
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


_OPEN_WORKBOOKS: list[Any] = []


@lru_cache(maxsize=16)
def _load_cached_workbook(path: str, size: int, modified_ns: int) -> Any:
    workbook = load_workbook(path, read_only=True, data_only=True)
    _OPEN_WORKBOOKS.append(workbook)
    return workbook


def load_read_only_workbook(path: Path) -> Any:
    """Reuse a workbook while its path, size and modification time are unchanged."""
    resolved = path.resolve()
    stat = resolved.stat()
    return _load_cached_workbook(str(resolved), stat.st_size, stat.st_mtime_ns)


def load_value_workbook(path: Path) -> Any:
    """Load export-style workbooks whose merged ranges fail in read-only mode."""
    return load_workbook(path.resolve(), read_only=False, data_only=True)


@atexit.register
def _close_cached_workbooks() -> None:
    for workbook in _OPEN_WORKBOOKS:
        try:
            workbook.close()
        except Exception:
            pass

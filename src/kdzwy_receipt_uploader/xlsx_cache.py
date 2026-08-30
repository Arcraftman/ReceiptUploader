"""Small, explicit Excel workbook loaders."""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def load_read_only_workbook(path: Path) -> Any:
    """Open a read-only workbook; the caller must close it."""
    return load_workbook(path.resolve(), read_only=True, data_only=True)


def load_value_workbook(path: Path) -> Any:
    """Load export-style workbooks whose merged ranges fail in read-only mode."""
    return load_workbook(path.resolve(), read_only=False, data_only=True)
